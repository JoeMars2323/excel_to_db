from openpyxl import load_workbook
from openpyxl.styles import Font

workbook = load_workbook('banking_marketing.xlsx')
sheet = workbook.active

sheet.insert_cols(idx=22)
sheet.font = Font(bold=True)

new_title_cell = sheet.cell(row=1, column=22)
new_title_cell.font = Font(bold=True)
new_title_cell.value = "new_value"

workbook.save('banking_marketing2.xlsx')